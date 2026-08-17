import numpy as np
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import date, datetime
import dateutil.parser as dateparser

from config import (
    EXCHANGERS,
    DATA_FILE,
    EVENT_FILE
)

def load_process_data(path):

    df = pd.read_csv(path)

    df = df.dropna(subset=df.columns[3:], how="all")

    df["datetime"] = pd.to_datetime(
        df["Date"].astype(str) + " " + df["Time"].astype(str)
    )

    df = df.sort_values("datetime")

    df.columns = (
        df.columns
            .str.replace("South Battery_Battery_Process Data_", "", regex=False)
            .str.replace("_Data_Process_Variable_Value", "_PV", regex=False)
            .str.replace("_Data_Control_Variable_Value", "_CV", regex=False)
            .str.replace("_Value", "", regex=False)
            .str.replace("-", "_", regex=False)
    )

    df = df.sort_values("Timestamp").reset_index(drop=True)

    df["datetime"] = pd.to_datetime(
        df["Date"].astype(str) + " " + df["Time"].astype(str)
    )

    df["date"] = df["datetime"].dt.date
    df = df.copy()
    return df


GREEN_RGB = "00B050"

EXCHANGER_COLUMNS = {
    "A": 2,
    "B": 3,
    "C": 4,
    "D": 5,
    "E": 6,
    "F": 7,
}


def cell_rgb(cell):

    fill = cell.fill

    if fill.patternType != "solid":
        return None

    if fill.fgColor.type == "rgb":
        return str(fill.fgColor.rgb).lstrip("#").upper()

    return None


def is_green(cell):

    rgb = cell_rgb(cell)

    if rgb is None:
        return False

    return rgb.endswith(GREEN_RGB)


def extract_serial(text):

    if text is None:
        return None

    text = str(text)

    m = re.search(r"#([\w-]+)", text)

    if m:
        return m.group(1)

    m = re.search(r"(\d{4,}-\d{4,})", text)

    return m.group(1) if m else None


def extract_clean_count(text):

    if text is None:
        return 0

    m = re.search(r"cleaned\s*\(?(\d+)", str(text), re.I)

    return int(m.group(1)) if m else 0


def parse_date(value):

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        return dateparser.parse(value.strip()).date()

    return value


def load_event_data(path):

    wb = load_workbook(path)

    ws = wb.active

    for sheet in wb.worksheets:
        if re.search(r"e560", sheet.title, re.I):
            ws = sheet
            break

    state = {
        x: {
            "serial": None,
            "serial_changed": False,
        }
        for x in EXCHANGER_COLUMNS
    }

    rows = []

    for excel_row in ws.iter_rows(min_row=2):

        date = excel_row[0].value

        if pd.isna(date):
            continue

        event = {"Date": parse_date(date)}

        for x, col in EXCHANGER_COLUMNS.items():

            cell = excel_row[col - 1]

            text = str(cell.value) if cell.value else ""

            serial = extract_serial(text)
            clean_count = extract_clean_count(text)

            if (
                serial is not None
                and state[x]["serial"] is not None
                and serial != state[x]["serial"]
            ):
                state[x]["serial_changed"] = True

            event[f"{x}_Clean"] = False
            event[f"{x}_Rebuild"] = False

            if is_green(cell):

                if state[x]["serial_changed"]:
                    event[f"{x}_Rebuild"] = True
                    state[x]["serial_changed"] = False

                elif clean_count > 0:
                    event[f"{x}_Clean"] = True

            if serial is not None:
                state[x]["serial"] = serial

        rows.append(event)

    return pd.DataFrame(rows)

def merge_events(df, events):

    for x in EXCHANGERS:
        # Create a set with dates and event types
        clean_dates = set(events.loc[events[f"{x}_Clean"] == True, "Date"])
        rebuild_dates = set(events.loc[events[f"{x}_Rebuild"] == True, "Date"])

        # Create event type columns in main dataframe and insert dates from earlier set.
        df[f"{x}_Clean"] = df["date"].isin(clean_dates)
        df[f"{x}_Rebuild"] = df["date"].isin(rebuild_dates)

    return df

def determine_online(df):
    for x in EXCHANGERS:

        fic = f"FIC_560{x}1_Data_Control Output"
        tic = f"TIC_560{x}2_Data_Control Output"

        df[f"{x}_Online"] = (
            (df[fic] > 0) &
            (df[tic] > 0)
        )
    df = df.copy()
    return df

def determine_cycles(df):
    
    cycle_columns = {}

    for x in EXCHANGERS:

        online = df[f"{x}_Online"]

        cycle_start = online & (~online.shift(fill_value=False))

        cycle = cycle_start.cumsum()

        cycle = cycle.where(online)

        valid_cycles = sorted(cycle.dropna().unique())

        mapping = {
            old: new
            for new, old in enumerate(valid_cycles[::-1], start=1)
        }

        cycle = cycle.map(mapping)

        cycle_start_time = (
            df.assign(temp_cycle=cycle)
            .groupby("temp_cycle")["datetime"]
            .transform("min")
        )

        cycle_columns[f"{x}_OperatingCycle"] = cycle

        cycle_columns[f"{x}_CycleStart"] = cycle_start_time

        cycle_columns[f"{x}_DaysInOperation"] = np.where(
            online,
            (df["datetime"] - cycle_start_time)
            .dt.total_seconds()
            / 86400,
            0.0
        )


    cycle_df = pd.DataFrame(cycle_columns)

    df = pd.concat(
        [
            df,
            cycle_df
        ],
        axis=1
    )
    df = df.copy()
    return df

def preprocess_data(df=None, events=None):

    if df is None:
        df = load_process_data(DATA_FILE)

    if events is None:
        events = load_event_data(EVENT_FILE)

    df = merge_events(df, events)
    df = determine_online(df)
    df = determine_cycles(df)

    return df
        
